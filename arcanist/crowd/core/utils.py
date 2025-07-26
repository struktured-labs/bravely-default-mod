import sys
import os
from pathlib import Path


# Required for pyinstaller
def get_filename(relative_path: str | Path) -> str:
    if os.path.exists(relative_path):
        filename = relative_path
    else:
        base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
        filename = os.path.join(base_path, relative_path)
    return filename.name if isinstance(filename, Path) else filename


import openpyxl
import xlwt

def xlsx_to_xls(xlsx_path:str|Path, xls_path:str|Path):
    # Load modern Excel file
    wb_xlsx = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Create legacy Excel file
    wb_xls = xlwt.Workbook()

    for sheetname in wb_xlsx.sheetnames:
        ws_xlsx = wb_xlsx[sheetname]
        ws_xls = wb_xls.add_sheet(sheetname[:31])  # .xls sheet name limit is 31 chars

        for row_idx, row in enumerate(ws_xlsx.iter_rows(values_only=True)):
            for col_idx, value in enumerate(row):
                if value is not None:
                    ws_xls.write(row_idx, col_idx, value)

    wb_xls.save(xls_path)
